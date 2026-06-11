"""Excel export — a tidy, sales-ready spreadsheet.

Produces a single formatted table: styled header, frozen header row, auto-filter,
sensible column widths, tier color coding, and CLICKABLE links (Website, WhatsApp,
Source URL) that open in the browser when clicked inside Excel.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (attribute, header, width). Order = column order in the sheet.
_COLUMNS = [
    ("company_name", "Company", 30),
    ("phone_e164", "Phone", 18),
    ("whatsapp", "WhatsApp", 16),
    ("extra_phones", "Extra Phones", 24),
    ("email", "Email", 28),
    ("website", "Website", 30),
    ("category", "Category", 18),
    ("company_type", "Type", 16),
    ("shipping_intent", "Intent", 8),
    ("target_markets", "Markets", 22),
    ("city", "City", 14),
    ("country", "Country", 12),
    ("address", "Address", 28),
    ("score", "Score", 7),
    ("tier", "Tier", 9),
    ("status", "Status", 14),
    ("assigned_to", "Assigned To", 14),
    ("next_followup_date", "Next Follow-up", 15),
    ("source", "Source", 16),
    ("sources_seen", "All Sources", 24),
    ("source_url", "Source URL", 30),
    # v2 — richer lead data (appended so existing columns keep their positions).
    ("segment", "Segment", 10),
    ("product_type", "Product Type", 18),
    ("store_platform", "Store", 12),
    ("facebook", "Facebook", 28),
    ("linkedin", "LinkedIn", 28),
    ("contact_name", "Contact Name", 18),
    ("contact_role", "Contact Role", 16),
    ("contact_email", "Contact Email", 24),
]

# Columns rendered as clickable hyperlinks: header -> URL builder.
_LINK_COLUMNS = {
    "Website": lambda v: v if str(v).startswith("http") else f"https://{v}",
    "Source URL": lambda v: v,
    "WhatsApp": lambda v: f"https://wa.me/{str(v).lstrip('+')}",
    "Facebook": lambda v: v if str(v).startswith("http") else f"https://{v}",
    "LinkedIn": lambda v: v if str(v).startswith("http") else f"https://{v}",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_LINK_FONT = Font(color="0563C1", underline="single")
_TIER_FILL = {
    "Hot": PatternFill("solid", fgColor="FFC7CE"),
    "Medium": PatternFill("solid", fgColor="FFEB9C"),
    "Weak": PatternFill("solid", fgColor="EDEDED"),
}


def _whatsapp_of(lead) -> str | None:
    links = getattr(lead, "social_links", None) or {}
    wa = links.get("whatsapp")
    if wa:
        return wa.rsplit("/", 1)[-1]  # the number part
    return None


def leads_to_dataframe(leads: Iterable) -> pd.DataFrame:
    rows = []
    for lead in leads:
        row = {}
        for attr, header, _ in _COLUMNS:
            if attr == "whatsapp":
                row[header] = _whatsapp_of(lead)
                continue
            val = getattr(lead, attr, None)
            if attr in ("extra_phones", "target_markets") and isinstance(val, list):
                val = ", ".join(val)
            row[header] = val
        rows.append(row)
    return pd.DataFrame(rows, columns=[h for _, h, _ in _COLUMNS])


def write_styled_workbook(df: pd.DataFrame, target) -> None:
    """Write a styled Leads sheet to a path or file-like buffer."""
    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        _style_sheet(writer.sheets["Leads"], df)


def export_excel(leads: Iterable, path: Path | str) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    write_styled_workbook(leads_to_dataframe(leads), path)
    return path


def _style_sheet(ws, df: pd.DataFrame) -> None:
    headers = list(df.columns)
    col_index = {h: i + 1 for i, h in enumerate(headers)}
    widths = {h: w for _, h, w in _COLUMNS}

    # Header row
    for h in headers:
        cell = ws.cell(row=1, column=col_index[h])
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    for h in headers:
        ws.column_dimensions[get_column_letter(col_index[h])].width = widths.get(h, 16)

    # Clickable links + tier coloring, row by row
    tier_col = col_index.get("Tier")
    for r in range(2, len(df) + 2):
        for header, build in _LINK_COLUMNS.items():
            ci = col_index.get(header)
            if not ci:
                continue
            cell = ws.cell(row=r, column=ci)
            val = cell.value
            if val not in (None, "", "None"):
                cell.hyperlink = build(val)
                cell.font = _LINK_FONT
        if tier_col:
            tval = ws.cell(row=r, column=tier_col).value
            fill = _TIER_FILL.get(tval)
            if fill:
                ws.cell(row=r, column=tier_col).fill = fill

    # Freeze header + enable auto-filter so the salesperson can sort/filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"

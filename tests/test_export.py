"""Styled Excel export: clickable links + tidy formatting."""
from openpyxl import load_workbook

from aol_leadfinder.export.excel import export_excel, leads_to_dataframe
from aol_leadfinder.storage.models import Lead


def _lead(**kw):
    base = dict(company_name="Acme", company_name_norm="acme")
    base.update(kw)
    return Lead(**base)


def test_dataframe_has_whatsapp_column():
    lead = _lead(social_links={"whatsapp": "https://wa.me/201001234567"})
    df = leads_to_dataframe([lead])
    assert "WhatsApp" in df.columns
    assert df.iloc[0]["WhatsApp"] == "201001234567"


def test_dataframe_surfaces_all_sources_provenance():
    lead = _lead(source="google_maps", sources_seen="egydir,google_maps")
    df = leads_to_dataframe([lead])
    assert "All Sources" in df.columns
    assert df.iloc[0]["All Sources"] == "egydir,google_maps"


def test_export_creates_clickable_links(tmp_path):
    lead = _lead(
        website="https://acme-eg.com",
        source_url="https://forwardingcompanies.com/company/acme",
        social_links={"whatsapp": "https://wa.me/201001234567"},
        tier="Hot",
        phone_e164="+201001234567",
    )
    out = export_excel([lead], tmp_path / "leads.xlsx")
    assert out.exists()

    wb = load_workbook(out)
    ws = wb["Leads"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    # Website / Source URL / WhatsApp cells carry a hyperlink
    assert ws.cell(row=2, column=col["Website"]).hyperlink is not None
    assert ws.cell(row=2, column=col["Source URL"]).hyperlink is not None
    assert ws.cell(row=2, column=col["WhatsApp"]).hyperlink is not None
    assert "wa.me" in ws.cell(row=2, column=col["WhatsApp"]).hyperlink.target

    # Header is frozen and auto-filter is enabled (tidy, sortable table)
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_website_without_scheme_gets_https(tmp_path):
    out = export_excel([_lead(website="acme-eg.com")], tmp_path / "l.xlsx")
    ws = load_workbook(out)["Leads"]
    headers = [c.value for c in ws[1]]
    ci = headers.index("Website") + 1
    assert ws.cell(row=2, column=ci).hyperlink.target.startswith("https://")
